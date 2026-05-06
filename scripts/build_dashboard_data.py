from __future__ import annotations

import hashlib
import json
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import geonamescache
import openpyxl


ROOT = Path(__file__).resolve().parent.parent
PUBLIC_JSON_OUTPUT_PATH = ROOT / "public" / "data" / "dashboard-data.json"
LEGACY_JS_OUTPUT_PATH = ROOT / "data" / "dashboard-data.js"
NOW = datetime.now().replace(microsecond=0)
COLOMBIA_CENTER = {"lat": 4.5709, "lon": -74.2973, "region": "Nacional"}
DEPARTMENT_GEO = {
    "AMAZONAS": {"lat": -1.54, "lon": -71.26, "region": "Amazonía"},
    "ANTIOQUIA": {"lat": 6.55, "lon": -75.83, "region": "Andes"},
    "BOLIVAR": {"lat": 8.24, "lon": -74.36, "region": "Caribe"},
    "BOYACA": {"lat": 5.45, "lon": -73.36, "region": "Andes"},
    "CALDAS": {"lat": 5.30, "lon": -75.25, "region": "Andes"},
    "CASANARE": {"lat": 5.54, "lon": -71.89, "region": "Orinoquía"},
    "CAUCA": {"lat": 2.57, "lon": -76.83, "region": "Pacífico"},
    "CESAR": {"lat": 9.34, "lon": -73.65, "region": "Caribe"},
    "CORDOBA": {"lat": 8.05, "lon": -75.58, "region": "Caribe"},
    "CUNDINAMARCA": {"lat": 4.86, "lon": -74.03, "region": "Andes"},
    "GUAINIA": {"lat": 2.58, "lon": -68.52, "region": "Amazonía"},
    "GUAVIARE": {"lat": 2.04, "lon": -72.33, "region": "Amazonía"},
    "LA GUAJIRA": {"lat": 11.35, "lon": -72.52, "region": "Caribe"},
    "MAGDALENA": {"lat": 10.41, "lon": -74.41, "region": "Caribe"},
    "META": {"lat": 3.27, "lon": -73.09, "region": "Orinoquía"},
    "NARINO": {"lat": 1.29, "lon": -77.36, "region": "Pacífico"},
    "NORTE DE SANTANDER": {"lat": 8.09, "lon": -72.91, "region": "Andes"},
    "PUTUMAYO": {"lat": 0.44, "lon": -76.52, "region": "Amazonía"},
    "QUINDIO": {"lat": 4.53, "lon": -75.67, "region": "Andes"},
    "RISARALDA": {"lat": 5.32, "lon": -75.93, "region": "Andes"},
    "SAN ANDRES": {"lat": 12.58, "lon": -81.70, "region": "Insular"},
    "SANTANDER": {"lat": 6.89, "lon": -73.37, "region": "Andes"},
    "SUCRE": {"lat": 9.30, "lon": -75.39, "region": "Caribe"},
    "TOLIMA": {"lat": 4.03, "lon": -75.26, "region": "Andes"},
    "VALLE DEL CAUCA": {"lat": 3.80, "lon": -76.64, "region": "Pacífico"},
    "VAUPES": {"lat": 0.86, "lon": -70.81, "region": "Amazonía"},
    "VICHADA": {"lat": 4.42, "lon": -69.29, "region": "Orinoquía"},
}
TOPIC_PATTERNS = {
    "Conectividad": ["trafico", "internet", "navegacion", "navegación", "wan", "ap interno", "ap externo", "sin servicio", "offline", "online", "senal", "señal"],
    "Energía": ["energia", "energía", "electrica", "eléctrica", "voltaje", "ups", "sin energia", "sin energía", "breaker"],
    "Infraestructura": ["antena", "cable", "router", "switch", "radio", "poste", "torre", "dano", "daño", "reubicacion", "reubicación"],
    "Acceso / Contacto": ["sin contacto", "sin respuesta", "no responde", "no contesta", "locatario", "visita", "sitio", "terreno", "desplazamiento", "ingreso"],
    "Mantenimiento": ["mantenimiento", "instalacion", "instalación", "ajuste", "alineacion", "alineación", "configuracion", "configuración"],
    "Administrativo": ["paz y salvo", "acta", "validacion", "validación", "cierre", "documentacion", "documentación"],
}
ALERT_LABELS = {
    "narrativa-sin-solucion": "Sin solución documentada",
    "narrativa-sin-contacto": "Dificultad de contacto",
    "narrativa-brecha-sla": "Brecha de SLA",
    "narrativa-energia": "Incidencia energética",
    "narrativa-infraestructura": "Infraestructura crítica",
    "narrativa-ticket-antiguo": "Ticket abierto envejecido",
    "narrativa-administrativo": "Cierre administrativo",
}


def main() -> None:
    source_path = resolve_source_path()
    workbook = openpyxl.load_workbook(source_path, read_only=True, data_only=True)

    ticket_sheet = workbook[workbook.sheetnames[0]]
    stop_sheet = workbook[workbook.sheetnames[1]]

    raw_tickets = list(iter_records(ticket_sheet))
    raw_stops = list(iter_records(stop_sheet))
    geo_registry = build_geo_registry(raw_tickets, raw_stops)

    stop_events, stop_aggregations = build_stop_payload(raw_stops, geo_registry)
    tickets = build_ticket_payload(raw_tickets, stop_aggregations, geo_registry)
    tickets, orphan_count = attach_orphan_stop_tickets(tickets, stop_aggregations, geo_registry)

    payload = {
        "meta": {
            "sourceFile": source_path.name,
            "generatedAt": NOW.isoformat(),
            "sheetNames": workbook.sheetnames,
            "recordCounts": {
                "tickets": len(tickets),
                "stopEvents": len(stop_events),
                "orphanStopTickets": orphan_count,
            },
            "coverage": {
                "clientCreatedPct": round(non_empty_ratio(raw_tickets, "Fecha Creacion Cliente") * 100, 2),
                "technicianPct": round(non_empty_ratio(raw_tickets, "Tecnico Asignado") * 100, 2),
                "maintenancePct": round(non_empty_ratio(raw_tickets, "Mantenimiento Id") * 100, 2),
                "resolutionCommentPct": round(non_empty_ratio(raw_tickets, "Comentario Solución") * 100, 2),
            },
            "geoCoverage": build_geo_coverage(tickets),
            "nlpSummary": build_nlp_summary(tickets),
            "warnings": build_warnings(raw_tickets, orphan_count, tickets),
        },
        "tickets": tickets,
        "stopEvents": stop_events,
    }

    PUBLIC_JSON_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    PUBLIC_JSON_OUTPUT_PATH.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")

    LEGACY_JS_OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    LEGACY_JS_OUTPUT_PATH.write_text(
        "window.__DASHBOARD_DATA__ = " + json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + ";\n",
        encoding="utf-8",
    )

    print(f"Dataset JSON generado en: {PUBLIC_JSON_OUTPUT_PATH}")
    print(f"Dataset legacy generado en: {LEGACY_JS_OUTPUT_PATH}")
    print(f"Fuente procesada: {source_path.name}")
    print(f"Tickets: {len(tickets)} | Eventos de parada: {len(stop_events)} | Tickets huérfanos integrados: {orphan_count}")


def resolve_source_path() -> Path:
    if len(sys.argv) > 1:
        candidate = Path(sys.argv[1]).expanduser().resolve()
        if not candidate.exists():
            raise FileNotFoundError(f"No existe el archivo indicado: {candidate}")
        return candidate

    env_source = os.environ.get("DASHBOARD_SOURCE_XLSX", "").strip()
    if env_source:
        candidate = Path(env_source).expanduser().resolve()
        if not candidate.exists():
            raise FileNotFoundError(f"No existe DASHBOARD_SOURCE_XLSX: {candidate}")
        return candidate

    raise FileNotFoundError(
        "No se definió el archivo fuente del dataset.\n"
        "Modo estricto activo: pasa el .xlsx por argumento o define DASHBOARD_SOURCE_XLSX.\n"
        "La fuente oficial debe venir de Drive."
    )


def iter_records(worksheet: Any):
    headers = list(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        record: dict[str, Any] = {}
        for index, header in enumerate(headers):
            if not header:
                continue
            value = row[index] if index < len(row) else None
            record[str(header)] = clean_value(value)
        if any(value not in (None, "") for value in record.values()):
            yield record


def build_geo_registry(raw_tickets: list[dict[str, Any]], raw_stops: list[dict[str, Any]]) -> dict[tuple[str, str], dict[str, Any]]:
    keys: set[tuple[str, str]] = set()
    for record in raw_tickets:
        keys.add((normalize_key(record.get("Departamento")), normalize_key(record.get("Municipio"))))
    for record in raw_stops:
        keys.add((normalize_key(record.get("Departamento")), normalize_key(record.get("Municipio"))))

    cache = geonamescache.GeonamesCache()
    city_index = build_city_index(cache)
    geo_registry: dict[tuple[str, str], dict[str, Any]] = {}

    for department_key, municipality_key in keys:
        if not department_key or not municipality_key:
            continue
        department_geo = DEPARTMENT_GEO.get(department_key, COLOMBIA_CENTER)
        exact_city = pick_city(city_index.get(municipality_key, []))
        if exact_city:
            lat = float(exact_city["latitude"])
            lon = float(exact_city["longitude"])
            precision = "municipality-exact"
        else:
            lat, lon = apply_jitter(department_geo["lat"], department_geo["lon"], municipality_key)
            precision = "department-approx"

        geo_registry[(department_key, municipality_key)] = {
            "lat": round(lat, 6),
            "lon": round(lon, 6),
            "precision": precision,
            "region": department_geo["region"],
            "departmentLat": round(float(department_geo["lat"]), 6),
            "departmentLon": round(float(department_geo["lon"]), 6),
        }

    return geo_registry


def build_city_index(cache: geonamescache.GeonamesCache) -> dict[str, list[dict[str, Any]]]:
    city_index: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for city in cache.get_cities().values():
        if city.get("countrycode") != "CO":
            continue
        names = {normalize_key(city.get("name"))}
        for alt_name in city.get("alternatenames") or []:
            normalized = normalize_key(alt_name)
            if normalized:
                names.add(normalized)
        for normalized_name in names:
            city_index[normalized_name].append(city)
    return city_index


def pick_city(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    return sorted(candidates, key=lambda item: item.get("population", 0), reverse=True)[0]


def apply_jitter(base_lat: float, base_lon: float, seed_text: str) -> tuple[float, float]:
    digest = hashlib.md5(seed_text.encode("utf-8")).digest()
    lat_offset = ((digest[0] / 255) - 0.5) * 0.9
    lon_offset = ((digest[1] / 255) - 0.5) * 1.1
    return base_lat + lat_offset, base_lon + lon_offset


def build_stop_payload(
    raw_stops: list[dict[str, Any]], geo_registry: dict[tuple[str, str], dict[str, Any]]
) -> tuple[list[dict[str, Any]], dict[str, dict[str, Any]]]:
    stop_events: list[dict[str, Any]] = []
    aggregations: dict[str, dict[str, Any]] = {}

    for record in raw_stops:
        ticket_id = stringify_id(record.get("Ticket #"))
        if not ticket_id:
            continue

        geo = resolve_geo_payload(record.get("Departamento"), record.get("Municipio"), geo_registry)
        event = {
            "ticketId": ticket_id,
            "operatorCode": clean_text(record.get("Código Operador")),
            "beneficiaryId": stringify_id(record.get("ID Beneficiario Mintic")),
            "department": clean_text(record.get("Departamento")),
            "municipality": clean_text(record.get("Municipio")),
            "locality": clean_text(record.get("Centro Poblado")),
            "periodLabel": clean_text(record.get("Año-Mes")),
            "faultStart": to_iso(record.get("Fecha Inicio Falla")),
            "faultEnd": to_iso(record.get("Fecha Fin Falla")),
            "stopStart": to_iso(record.get("Fecha Inicio Parada")),
            "stopEnd": to_iso(record.get("Fecha Fin Parada")),
            "stopDays": round_number(to_number(record.get("Días Parada Reloj")), 4),
            "geo": geo,
        }
        stop_events.append(event)

        current = aggregations.setdefault(
            ticket_id,
            {
                "ticketId": ticket_id,
                "operatorCode": event["operatorCode"],
                "beneficiaryId": event["beneficiaryId"],
                "department": event["department"],
                "municipality": event["municipality"],
                "locality": event["locality"],
                "faultStart": event["faultStart"],
                "faultEnd": event["faultEnd"],
                "stopStart": event["stopStart"],
                "stopEnd": event["stopEnd"],
                "stopDaysTotal": 0.0,
                "stopSegmentCount": 0,
                "maxStopDays": 0.0,
                "geo": geo,
            },
        )

        current["stopDaysTotal"] = round_number((current["stopDaysTotal"] or 0) + (event["stopDays"] or 0), 4)
        current["stopSegmentCount"] += 1
        current["maxStopDays"] = max(current["maxStopDays"], event["stopDays"] or 0)
        current["faultStart"] = min_iso(current["faultStart"], event["faultStart"])
        current["faultEnd"] = max_iso(current["faultEnd"], event["faultEnd"])
        current["stopStart"] = min_iso(current["stopStart"], event["stopStart"])
        current["stopEnd"] = max_iso(current["stopEnd"], event["stopEnd"])

    for aggregation in aggregations.values():
        aggregation["faultDurationDays"] = round_number(duration_days(aggregation.get("faultStart"), aggregation.get("faultEnd")), 4)

    return stop_events, aggregations


def build_ticket_payload(
    raw_tickets: list[dict[str, Any]],
    stop_aggregations: dict[str, dict[str, Any]],
    geo_registry: dict[tuple[str, str], dict[str, Any]],
) -> list[dict[str, Any]]:
    tickets: list[dict[str, Any]] = []

    for record in raw_tickets:
        ticket_id = stringify_id(record.get("#Ticket"))
        if not ticket_id:
            continue

        ticket_start = to_iso(record.get("Fecha Inicio"))
        ticket_end = to_iso(record.get("Ticket Fecha Fin"))
        client_created = to_iso(record.get("Fecha Creacion Cliente"))
        stop_data = stop_aggregations.get(ticket_id, {})
        geo = resolve_geo_payload(record.get("Departamento"), record.get("Municipio"), geo_registry)

        ticket = {
            "ticketId": ticket_id,
            "operatorCode": clean_text(record.get("Código Operador")),
            "clientCode": stringify_id(record.get("Código Cliente")),
            "department": clean_text(record.get("Departamento")),
            "municipality": clean_text(record.get("Municipio")),
            "locality": clean_text(record.get("Centro Poblado")),
            "ticketStart": ticket_start,
            "ticketEnd": ticket_end,
            "referenceDate": client_created or ticket_start,
            "ticketState": clean_text(record.get("Ticket Estado")) or "Sin dato",
            "creator": clean_text(record.get("Usuario Crea")),
            "assignee": clean_text(record.get("Asignado A")),
            "closer": clean_text(record.get("Usuario Cierre")),
            "maintenanceId": stringify_id(record.get("Mantenimiento Id")),
            "maintenanceStart": to_iso(record.get("Fecha Inicio Mantenimiento")),
            "technician": clean_text(record.get("Tecnico Asignado")),
            "maintenanceCreator": clean_text(record.get("Usuario Crea Mnt")),
            "maintenanceEnd": to_iso(record.get("Fecha Fin Mnt")),
            "maintenanceCloser": clean_text(record.get("Usuario Cierre Mnt")),
            "maintenanceState": clean_text(record.get("Estado Mnt")),
            "category": clean_text(record.get("Categoria")),
            "sourceOrigin": clean_text(record.get("Fuente Origen")),
            "escalationGroup": clean_text(record.get("Grupo Escalamiento")),
            "impact": clean_text(record.get("Impacto")),
            "urgency": clean_text(record.get("Urgencia")),
            "priority": clean_text(record.get("Prioridad")) or "Sin dato",
            "subProject": clean_text(record.get("Sub Proyecto")),
            "project": clean_text(record.get("Proyecto")),
            "type": clean_text(record.get("Tipo")),
            "responsible": clean_text(record.get("Responsable")),
            "clientCreated": client_created,
            "infoSource": clean_text(record.get("Fuente Información")),
            "openingComment": clean_text(record.get("Comentario Apertura")),
            "resolutionComment": clean_text(record.get("Comentario Solución")),
            "beneficiaryId": stop_data.get("beneficiaryId"),
            "stopDaysTotal": round_number(stop_data.get("stopDaysTotal"), 4),
            "stopSegmentCount": stop_data.get("stopSegmentCount", 0),
            "stopStart": stop_data.get("stopStart"),
            "stopEnd": stop_data.get("stopEnd"),
            "faultStart": stop_data.get("faultStart"),
            "faultEnd": stop_data.get("faultEnd"),
            "faultDurationDays": round_number(stop_data.get("faultDurationDays"), 4),
            "closeDurationDays": round_number(duration_days(ticket_start, ticket_end), 4),
            "openAgeDays": round_number(duration_days(ticket_start, NOW.isoformat()), 4) if clean_text(record.get("Ticket Estado")) == "Abierto" else None,
            "geo": geo,
        }
        ticket["textAnalytics"] = build_text_analytics(ticket)
        tickets.append(ticket)

    return tickets


def attach_orphan_stop_tickets(
    tickets: list[dict[str, Any]],
    stop_aggregations: dict[str, dict[str, Any]],
    geo_registry: dict[tuple[str, str], dict[str, Any]],
) -> tuple[list[dict[str, Any]], int]:
    known_ticket_ids = {ticket["ticketId"] for ticket in tickets}
    orphan_count = 0

    for ticket_id, stop_data in stop_aggregations.items():
        if ticket_id in known_ticket_ids:
            continue

        orphan_count += 1
        ticket = {
            "ticketId": ticket_id,
            "operatorCode": stop_data.get("operatorCode"),
            "clientCode": None,
            "department": stop_data.get("department"),
            "municipality": stop_data.get("municipality"),
            "locality": stop_data.get("locality"),
            "ticketStart": stop_data.get("faultStart"),
            "ticketEnd": stop_data.get("faultEnd"),
            "referenceDate": stop_data.get("faultStart") or stop_data.get("stopStart"),
            "ticketState": "Sin dato",
            "creator": None,
            "assignee": None,
            "closer": None,
            "maintenanceId": None,
            "maintenanceStart": None,
            "technician": None,
            "maintenanceCreator": None,
            "maintenanceEnd": None,
            "maintenanceCloser": None,
            "maintenanceState": None,
            "category": None,
            "sourceOrigin": None,
            "escalationGroup": None,
            "impact": None,
            "urgency": None,
            "priority": "Sin dato",
            "subProject": None,
            "project": "Solo paradas",
            "type": None,
            "responsible": None,
            "clientCreated": None,
            "infoSource": None,
            "openingComment": None,
            "resolutionComment": None,
            "beneficiaryId": stop_data.get("beneficiaryId"),
            "stopDaysTotal": round_number(stop_data.get("stopDaysTotal"), 4),
            "stopSegmentCount": stop_data.get("stopSegmentCount", 0),
            "stopStart": stop_data.get("stopStart"),
            "stopEnd": stop_data.get("stopEnd"),
            "faultStart": stop_data.get("faultStart"),
            "faultEnd": stop_data.get("faultEnd"),
            "faultDurationDays": round_number(stop_data.get("faultDurationDays"), 4),
            "closeDurationDays": round_number(duration_days(stop_data.get("faultStart"), stop_data.get("faultEnd")), 4),
            "openAgeDays": None,
            "geo": stop_data.get("geo") or resolve_geo_payload(stop_data.get("department"), stop_data.get("municipality"), geo_registry),
        }
        ticket["textAnalytics"] = build_text_analytics(ticket)
        tickets.append(ticket)

    tickets.sort(key=lambda item: (item.get("referenceDate") or "", item.get("ticketId") or ""))
    return tickets, orphan_count


def build_text_analytics(ticket: dict[str, Any]) -> dict[str, Any]:
    normalized = normalize_text(
        " ".join(
            filter(
                None,
                [
                    ticket.get("category"),
                    ticket.get("type"),
                    ticket.get("openingComment"),
                    ticket.get("resolutionComment"),
                ],
            )
        )
    )

    topic_scores: dict[str, int] = {}
    matched_signals: list[str] = []
    for topic, phrases in TOPIC_PATTERNS.items():
        score = 0
        for phrase in phrases:
            normalized_phrase = normalize_text(phrase)
            occurrences = normalized.count(normalized_phrase)
            if occurrences:
                score += occurrences * max(1, len(normalized_phrase.split()))
                matched_signals.append(phrase)
        topic_scores[topic] = score

    total_score = sum(topic_scores.values())
    primary_topic = max(topic_scores, key=topic_scores.get) if total_score else "Sin clasificar"
    confidence = round(topic_scores.get(primary_topic, 0) / total_score, 3) if total_score else 0.0

    alerts: list[str] = []
    if not clean_text(ticket.get("resolutionComment")):
        alerts.append("narrativa-sin-solucion")
    if (ticket.get("stopDaysTotal") or 0) > 7:
        alerts.append("narrativa-brecha-sla")
    if contains_any(normalized, ["sin contacto", "sin respuesta", "no responde", "no contesta"]):
        alerts.append("narrativa-sin-contacto")
    if topic_scores.get("Energía", 0) > 0:
        alerts.append("narrativa-energia")
    if topic_scores.get("Infraestructura", 0) > 0 and ticket.get("ticketState") == "Abierto":
        alerts.append("narrativa-infraestructura")
    if (ticket.get("openAgeDays") or 0) > 30:
        alerts.append("narrativa-ticket-antiguo")
    if topic_scores.get("Administrativo", 0) > 0:
        alerts.append("narrativa-administrativo")

    return {
        "primaryTopic": primary_topic,
        "confidence": confidence,
        "topicScores": {topic: score for topic, score in topic_scores.items() if score > 0},
        "alerts": alerts,
        "alertLabels": [ALERT_LABELS[alert] for alert in alerts if alert in ALERT_LABELS],
        "matchedSignals": sorted(set(normalize_text(signal).title() for signal in matched_signals))[:8],
        "resolutionDocumented": bool(clean_text(ticket.get("resolutionComment"))),
    }


def build_geo_coverage(tickets: list[dict[str, Any]]) -> dict[str, int]:
    counter = Counter(ticket.get("geo", {}).get("precision", "unknown") for ticket in tickets)
    return dict(counter)


def build_nlp_summary(tickets: list[dict[str, Any]]) -> dict[str, dict[str, int]]:
    topic_counter = Counter(ticket.get("textAnalytics", {}).get("primaryTopic", "Sin clasificar") for ticket in tickets)
    alert_counter = Counter()
    for ticket in tickets:
        for alert in ticket.get("textAnalytics", {}).get("alerts", []):
            alert_counter[ALERT_LABELS.get(alert, alert)] += 1
    return {"topics": dict(topic_counter), "alerts": dict(alert_counter)}


def resolve_geo_payload(
    department: Any,
    municipality: Any,
    geo_registry: dict[tuple[str, str], dict[str, Any]],
) -> dict[str, Any]:
    department_key = normalize_key(department)
    municipality_key = normalize_key(municipality)
    department_geo = DEPARTMENT_GEO.get(department_key, COLOMBIA_CENTER)
    resolved = geo_registry.get((department_key, municipality_key))
    if resolved:
        return {
            "departmentKey": department_key,
            "municipalityKey": municipality_key,
            **resolved,
        }
    return {
        "departmentKey": department_key,
        "municipalityKey": municipality_key,
        "lat": department_geo["lat"],
        "lon": department_geo["lon"],
        "precision": "department-fallback",
        "region": department_geo["region"],
        "departmentLat": department_geo["lat"],
        "departmentLon": department_geo["lon"],
    }


def build_warnings(raw_tickets: list[dict[str, Any]], orphan_count: int, tickets: list[dict[str, Any]]) -> list[str]:
    warnings: list[str] = []

    if non_empty_ratio(raw_tickets, "Fecha Creacion Cliente") == 0:
        warnings.append(
            "Fecha Creacion Cliente viene vacía en todos los registros; el dashboard usa Fecha Inicio como referencia principal de creación."
        )

    if non_empty_ratio(raw_tickets, "Tecnico Asignado") < 0.05:
        warnings.append(
            "Tecnico Asignado tiene baja cobertura; la carga operativa se interpreta principalmente con Asignado A y Grupo Escalamiento."
        )

    if orphan_count:
        warnings.append(
            f"Se integraron {orphan_count} tickets desde la hoja de paradas porque no estaban presentes en la hoja principal de tickets."
        )

    approx_count = sum(1 for ticket in tickets if ticket.get("geo", {}).get("precision") != "municipality-exact")
    if approx_count:
        warnings.append(
            f"{approx_count} tickets usan coordenadas aproximadas a nivel departamental cuando no hubo match municipal exacto en la base geográfica local."
        )

    warnings.append("Las métricas de parada se agregan por ticket a partir de la hoja DetParadas.")
    warnings.append("La capa NLP usa clasificación temática por lexicones y alertas narrativas derivadas del comentariado operativo.")
    return warnings


def clean_value(value: Any) -> Any:
    if isinstance(value, str):
        cleaned = value.replace("\r\n", "\n").strip()
        return cleaned or None
    return value


def clean_text(value: Any) -> str | None:
    if value is None:
        return None
    return str(value).strip() or None


def stringify_id(value: Any) -> str | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return str(value)
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        return str(int(value)) if value.is_integer() else str(value)
    text = str(value).strip()
    return text or None


def to_datetime(value: Any) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        formats = (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%dT%H:%M:%S",
            "%Y-%m-%d",
            "%d/%m/%Y %H:%M:%S",
            "%d/%m/%Y",
        )
        for fmt in formats:
            try:
                return datetime.strptime(value, fmt)
            except ValueError:
                continue
    return None


def to_iso(value: Any) -> str | None:
    date_value = to_datetime(value)
    return date_value.isoformat() if date_value else None


def to_number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).replace(",", ".").strip()
    try:
        return float(text)
    except ValueError:
        return None


def duration_days(start: str | None, end: str | None) -> float | None:
    start_date = to_datetime(start)
    end_date = to_datetime(end)
    if not start_date or not end_date or end_date < start_date:
        return None
    return (end_date - start_date).total_seconds() / 86400


def min_iso(left: str | None, right: str | None) -> str | None:
    if not left:
        return right
    if not right:
        return left
    return left if to_datetime(left) <= to_datetime(right) else right


def max_iso(left: str | None, right: str | None) -> str | None:
    if not left:
        return right
    if not right:
        return left
    return left if to_datetime(left) >= to_datetime(right) else right


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", str(value or "")).encode("ascii", "ignore").decode("ascii")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def normalize_key(value: Any) -> str:
    return normalize_text(value).upper()


def contains_any(text: str, phrases: list[str]) -> bool:
    return any(normalize_text(phrase) in text for phrase in phrases)


def non_empty_ratio(records: list[dict[str, Any]], field: str) -> float:
    if not records:
        return 0.0
    hits = sum(1 for record in records if record.get(field) not in (None, ""))
    return hits / len(records)


def round_number(value: float | None, digits: int) -> float | None:
    if value is None:
        return None
    return round(value, digits)


if __name__ == "__main__":
    main()