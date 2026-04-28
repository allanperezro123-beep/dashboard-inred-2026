#!/usr/bin/env python3
"""
preparar_juntas.py — Consolida los 4 archivos JUNTAS mensuales (DATOS HOY/) en:
  1. datos/JUNTAS_Consolidado_2026__<ts>.xlsx  → Excel que el dashboard puede leer directamente
  2. INFORME_JUNTAS_2026__<ts>.xlsx            → Informe legible multi-hoja para consulta humana

Uso:
  python scripts/preparar_juntas.py
  python scripts/preparar_juntas.py "DATOS HOY"   # carpeta personalizada
"""
from __future__ import annotations

import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT       = Path(__file__).resolve().parent.parent
SOURCE_DIR = ROOT / "DATOS HOY"
DATOS_DIR  = ROOT / "datos"

# ── Paleta de colores ─────────────────────────────────────────────────────────
C_DARK     = "0F2A4A"
C_MED      = "1B3A5C"
C_BLUE     = "2E6DA4"
C_ACCENT   = "E8F4FD"
C_WHITE    = "FFFFFF"
C_GREEN    = "D4EDDA"
C_GREEN2   = "DCF1E3"
C_YELLOW   = "FFF3CD"
C_YELLOW2  = "FFFDE7"
C_RED      = "F8D7DA"
C_RED2     = "FCEAEA"

# ── Columnas que espera build_dashboard_data.py  ─────────────────────────────
DASHBOARD_TICKET_COLS = [
    "Código Operador", "Código Cliente", "Departamento", "Municipio",
    "Centro Poblado", "#Ticket", "Fecha Inicio", "Usuario Crea",
    "Asignado A", "Ticket Fecha Fin", "Ticket Estado", "Usuario Cierre",
    "Mantenimiento Id", "Fecha Inicio Mantenimiento", "Tecnico Asignado",
    "Usuario Crea Mnt", "Fecha Fin Mnt", "Usuario Cierre Mnt", "Estado Mnt",
    "Categoria", "Fuente Origen", "Grupo Escalamiento", "Impacto", "Urgencia",
    "Prioridad", "Sub Proyecto", "Proyecto", "Tipo", "Responsable",
    "Fecha Creacion Cliente", "Fuente Información", "Comentario Apertura",
    "Comentario Solución",
]

DASHBOARD_STOP_COLS = [
    "Año-Mes", "Código Operador", "ID Beneficiario Mintic",
    "Departamento", "Municipio", "Centro Poblado", "Ticket #",
    "Fecha Inicio Falla", "Fecha Fin Falla",
    "Fecha Inicio Parada", "Fecha Fin Parada", "Días Parada Reloj",
]

# Mapeo columna-fuente → columna-dashboard para tickets JUNTAS
TICKET_TO_DASHBOARD = {
    "Código Operador":      "Código Operador",
    "ID Beneficiario Mintic": "Código Cliente",
    "Departamento":         "Departamento",
    "Municipio":            "Municipio",
    "Centro Poblado":       "Centro Poblado",
    "#Ticket":              "#Ticket",
    "Fecha Inicio":         "Fecha Inicio",
    "Fecha Cierre":         "Ticket Fecha Fin",
    "Estado":               "Ticket Estado",
    "Categoría":            "Categoria",
    "Prioridad":            "Prioridad",
    "Responsabilidad":      "Responsable",
    "Descripción Apertura": "Comentario Apertura",
    "Solución":             "Comentario Solución",
}


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    source_dir = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else SOURCE_DIR
    if not source_dir.is_dir():
        sys.exit(f"[ERROR] Carpeta no encontrada: {source_dir}")

    files = sorted(source_dir.glob("*.xlsx"), key=lambda p: p.name)
    if not files:
        sys.exit(f"[ERROR] No hay archivos .xlsx en: {source_dir}")

    print(f"\nProcesando {len(files)} archivos desde: {source_dir.name}/")

    all_tickets: list[dict] = []
    all_stops:   list[dict] = []
    all_disp:    list[dict] = []

    for filepath in files:
        print(f"  Leyendo: {filepath.name}")
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        snames = wb.sheetnames

        tix = read_sheet(wb[snames[0]])
        mes = extract_month(snames[0])
        for row in tix:
            row["__mes__"] = mes
        all_tickets.extend(tix)

        all_stops.extend(read_sheet(wb[snames[1]]))

        if len(snames) >= 3:
            all_disp.extend(read_sheet(wb[snames[2]]))

        wb.close()

    # Deduplicar tickets: si el mismo #Ticket aparece en varios meses, quedarse con el más reciente
    seen: dict[str, dict] = {}
    for row in all_tickets:
        tid = str(row.get("#Ticket") or "").strip()
        if tid:
            seen[tid] = row
    tickets = sorted(seen.values(),
                     key=lambda r: (str(r.get("Fecha Inicio") or ""), str(r.get("#Ticket") or "")))

    # Deduplicar paradas por (Ticket # + Fecha Inicio Parada)
    seen_s: dict[tuple, dict] = {}
    for row in all_stops:
        key = (str(row.get("Ticket #") or ""), str(row.get("Fecha Inicio Parada") or ""))
        seen_s[key] = row
    stops = sorted(seen_s.values(),
                   key=lambda r: (str(r.get("Año-Mes") or ""), str(r.get("Ticket #") or "")))

    print(f"\nTotal consolidado: {len(tickets)} tickets | {len(stops)} paradas | {len(all_disp)} registros disponibilidad")

    DATOS_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y-%m-%d %H_%M_%S")

    # ── Salida 1: Excel para el dashboard ────────────────────────────────────
    dash_path = DATOS_DIR / f"JUNTAS_Consolidado_2026__{ts}.xlsx"
    write_dashboard_excel(tickets, stops, dash_path, ts)
    print(f"\n[1] Dashboard Excel  →  {dash_path.relative_to(ROOT)}")

    # ── Salida 2: Informe legible ─────────────────────────────────────────────
    report_path = ROOT / f"INFORME_JUNTAS_2026__{ts}.xlsx"
    write_report_excel(tickets, stops, all_disp, report_path)
    print(f"[2] Informe legible  →  {report_path.name}")
    print("\nListo.\n")


# ── Lectura de hojas ──────────────────────────────────────────────────────────

def read_sheet(ws) -> list[dict]:
    rows_iter = ws.iter_rows(values_only=True)
    try:
        raw_headers = next(rows_iter)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in raw_headers]
    records = []
    for row in rows_iter:
        record = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            val = row[i] if i < len(row) else None
            record[h] = _clean(val)
        if any(v not in (None, "") for v in record.values()):
            records.append(record)
    return records


def _clean(v):
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def extract_month(sheet_name: str) -> str:
    """'DetalleTickets_2026-3'  →  '2026-03'"""
    parts = sheet_name.rsplit("-", 1)
    if len(parts) == 2 and parts[1].isdigit():
        return f"2026-{int(parts[1]):02d}"
    return sheet_name


def hours_to_days(val) -> float | None:
    if val is None:
        return None
    try:
        return round(float(val) / 24, 4)
    except (TypeError, ValueError):
        return None


def fmt_date(v) -> str:
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.strftime("%d/%m/%Y %H:%M")
    s = str(v).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d",
                "%d/%m/%Y %H:%M:%S", "%d/%m/%Y"):
        try:
            dt = datetime.strptime(s, fmt)
            return dt.strftime("%d/%m/%Y %H:%M") if (" " in fmt or "T" in fmt) else dt.strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s


# ── Escritura Excel para dashboard ───────────────────────────────────────────

def write_dashboard_excel(tickets: list[dict], stops: list[dict], path: Path, ts: str) -> None:
    wb = openpyxl.Workbook()
    date_tag = ts[:10]

    # Hoja 1: tickets
    ws1 = wb.active
    ws1.title = f"DetTicketsJuntas_{date_tag}"
    ws1.append(DASHBOARD_TICKET_COLS)

    inv_map = {}  # dashboard_col → source_col
    for src_col, dst_col in TICKET_TO_DASHBOARD.items():
        inv_map[dst_col] = src_col

    for src in tickets:
        row_out = []
        for col in DASHBOARD_TICKET_COLS:
            val = src.get(col)                       # direct match (unlikely)
            if val is None:
                src_col = inv_map.get(col)
                if src_col:
                    val = src.get(src_col)
            row_out.append(val)
        ws1.append(row_out)

    # Hoja 2: paradas
    ws2 = wb.create_sheet(f"DetParadasJuntas_{date_tag}")
    ws2.append(DASHBOARD_STOP_COLS)

    for src in stops:
        row_out = []
        for col in DASHBOARD_STOP_COLS:
            if col == "Días Parada Reloj":
                val = hours_to_days(src.get("Horas Parada Reloj"))
            else:
                val = src.get(col)
            row_out.append(val)
        ws2.append(row_out)

    wb.save(path)


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def hdr_style(ws, row_idx: int, bg: str, fg: str = "FFFFFF", size: int = 10) -> None:
    fill  = PatternFill("solid", fgColor=bg)
    font  = Font(bold=True, color=fg, size=size)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    bdr   = Border(bottom=Side(style="medium", color="888888"))
    for cell in ws[row_idx]:
        if cell.value is not None:
            cell.fill  = fill
            cell.font  = font
            cell.alignment = align
            cell.border = bdr


def row_fill(ws, row_idx: int, color: str, n_cols: int) -> None:
    f = PatternFill("solid", fgColor=color)
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = f
        cell.alignment = Alignment(vertical="center")


def col_widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title_row(ws, text: str, span: str, row: int = 1) -> None:
    ws.merge_cells(span)
    cell = ws[span.split(":")[0]]
    cell.value = text
    cell.font  = Font(bold=True, size=13, color="FFFFFF")
    cell.fill  = PatternFill("solid", fgColor=C_DARK)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[row].height = 30


def freeze_filter(ws, cell: str = "A3") -> None:
    ws.freeze_panes = cell
    ws.auto_filter.ref = ws.dimensions


# ── Escritura informe legible ─────────────────────────────────────────────────

def write_report_excel(tickets: list[dict], stops: list[dict], disp: list[dict], path: Path) -> None:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "1. Resumen General"
    _sheet_resumen(ws1, tickets, stops, disp)

    ws2 = wb.create_sheet("2. Tickets Completos")
    _sheet_tickets(ws2, tickets, "TICKETS CONSOLIDADOS — ENERO A ABRIL 2026")

    ws3 = wb.create_sheet("3. Paradas Detalle")
    _sheet_paradas(ws3, stops)

    ws4 = wb.create_sheet("4. Disponibilidad x Juntas")
    _sheet_disp(ws4, disp)

    ws5 = wb.create_sheet("5. Por Operador")
    _sheet_operador(ws5, tickets, stops, disp)

    ws6 = wb.create_sheet("6. Por Departamento")
    _sheet_departamento(ws6, tickets, stops)

    abiertos = [t for t in tickets
                if str(t.get("Estado") or "").strip().lower()
                not in ("cerrado", "closed", "resuelto")]
    ws7 = wb.create_sheet("7. Tickets Abiertos")
    _sheet_tickets(ws7, abiertos, "TICKETS ABIERTOS / SIN CERRAR")

    ws8 = wb.create_sheet("8. Cumplimiento SLA")
    _sheet_sla(ws8, tickets, stops)

    wb.save(path)


# ── Hoja 1: Resumen General ───────────────────────────────────────────────────

def _sheet_resumen(ws, tickets: list, stops: list, disp: list) -> None:
    now_str = datetime.now().strftime("%d/%m/%Y %H:%M")

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = "INFORME CONSOLIDADO JUNTAS 2026 — INRED"
    c.font  = Font(bold=True, size=16, color="FFFFFF")
    c.fill  = PatternFill("solid", fgColor=C_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    c2 = ws["A2"]
    c2.value = f"Generado: {now_str}   |   Cobertura: Enero – Abril 2026   |   Fuente: DATOS HOY/"
    c2.font  = Font(size=10, color="555555", italic=True)
    c2.alignment = Alignment(horizontal="center")

    ws.append([])

    # ── KPIs ──────────────────────────────────────────────────────────────────
    total_tix = len(tickets)
    cerrados  = sum(1 for t in tickets
                    if str(t.get("Estado") or "").strip().lower() in ("cerrado", "closed", "resuelto"))
    abiertos  = total_tix - cerrados
    total_stops   = len(stops)
    total_horas   = sum(float(s.get("Horas Parada Reloj") or 0) for s in stops)
    total_dias    = total_horas / 24
    n_operators   = len(set(str(t.get("Código Operador") or "").strip() for t in tickets if t.get("Código Operador")))
    n_depts       = len(set(str(t.get("Departamento") or "").strip() for t in tickets if t.get("Departamento")))
    n_munis       = len(set(str(t.get("Municipio") or "").strip() for t in tickets if t.get("Municipio")))
    pct_cierre    = f"{cerrados / total_tix * 100:.1f}%" if total_tix else "N/A"

    kpis = [
        ("Total de Tickets",              total_tix),
        ("Tickets Cerrados",              cerrados),
        ("Tickets Abiertos / Pendientes", abiertos),
        ("% de Cierre",                   pct_cierre),
        ("Total Eventos de Parada",       total_stops),
        ("Total Horas de Parada",         f"{total_horas:,.1f} h"),
        ("Total Días de Parada",          f"{total_dias:,.1f} días"),
        ("Operadores con tickets",        n_operators),
        ("Departamentos afectados",       n_depts),
        ("Municipios afectados",          n_munis),
        ("Registros Disponibilidad",      len(disp)),
    ]

    ws.append(["INDICADOR GLOBAL", "VALOR"])
    hdr_row = ws.max_row
    hdr_style(ws, hdr_row, C_MED)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 22

    for i, (label, val) in enumerate(kpis):
        ws.append([label, val])
        r = ws.max_row
        fgc = C_ACCENT if i % 2 == 0 else C_WHITE
        for col in ("A", "B"):
            cell = ws[f"{col}{r}"]
            cell.fill  = PatternFill("solid", fgColor=fgc)
            cell.alignment = Alignment(vertical="center")
            cell.font  = Font(size=10)
        ws[f"A{r}"].font = Font(bold=True, size=10)

    ws.append([])

    # ── Resumen por mes ───────────────────────────────────────────────────────
    ws.append(["MES", "TICKETS", "ABIERTOS", "CERRADOS", "% CIERRE",
               "PARADAS", "HORAS PARADA", "DÍAS PARADA"])
    hdr_row2 = ws.max_row
    hdr_style(ws, hdr_row2, C_BLUE)
    for col in ("A","B","C","D","E","F","G","H"):
        ws.column_dimensions[col].width = max(ws.column_dimensions[col].width, 14)

    MES_LABELS = {
        "2026-01": "Enero 2026", "2026-02": "Febrero 2026",
        "2026-03": "Marzo 2026", "2026-04": "Abril 2026",
    }
    for i, (m, label) in enumerate(MES_LABELS.items()):
        m_tix = [t for t in tickets if str(t.get("__mes__") or "") == m]
        m_num = str(int(m.split("-")[1]))
        # Año-Mes en fuente puede ser "2026-1", "2026-01", "2026-1 " etc.
        possible = {m, m.replace("-0", "-"), f"2026-{m_num}"}
        m_stops = [s for s in stops if str(s.get("Año-Mes") or "").strip() in possible]
        m_horas = sum(float(s.get("Horas Parada Reloj") or 0) for s in m_stops)
        m_dias  = m_horas / 24
        m_cerr  = sum(1 for t in m_tix
                      if str(t.get("Estado") or "").strip().lower()
                      in ("cerrado", "closed", "resuelto"))
        m_ab    = len(m_tix) - m_cerr
        m_pct   = f"{m_cerr / len(m_tix) * 100:.1f}%" if m_tix else "N/A"

        ws.append([label, len(m_tix), m_ab, m_cerr, m_pct,
                   len(m_stops), round(m_horas, 1), round(m_dias, 2)])
        r = ws.max_row
        fgc = C_ACCENT if i % 2 == 0 else C_WHITE
        row_fill(ws, r, fgc, 8)

    ws.freeze_panes = "A4"


# ── Hoja 2 y 7: Tickets ───────────────────────────────────────────────────────

TICKET_COLS = [
    ("Mes",                    12),
    ("#Ticket",                12),
    ("#Ticket CCC",            14),
    ("Código Operador",        22),
    ("ID Beneficiario Mintic", 20),
    ("ID Sede",                14),
    ("Departamento",           20),
    ("Municipio",              22),
    ("Centro Poblado",         26),
    ("Sede Principal",         30),
    ("DANE Sede Ppal.",        16),
    ("Sede a Beneficiar",      30),
    ("DANE Sede Benef.",       16),
    ("Tipo Sol. Energía",      18),
    ("Fecha Inicio",           17),
    ("Fecha Cierre",           17),
    ("Estado",                 13),
    ("Prioridad",              13),
    ("Afecta Servicio",        14),
    ("Responsabilidad",        22),
    ("Categoría",              22),
    ("Días Háb. Solución",     16),
    ("Días Cal. Solución",     16),
    ("Tiempo Paradas Reloj",   18),
    ("Descripción Apertura",   44),
    ("Solución",               44),
]


def _get_ticket_row(ticket: dict) -> list:
    return [
        ticket.get("__mes__"),
        ticket.get("#Ticket"),
        ticket.get("#Ticket CCC"),
        ticket.get("Código Operador"),
        ticket.get("ID Beneficiario Mintic"),
        ticket.get("ID Sede"),
        ticket.get("Departamento"),
        ticket.get("Municipio"),
        ticket.get("Centro Poblado"),
        ticket.get("Nombre Sede Educativa Principal"),
        ticket.get("Código Dane Sede Educativa Principal"),
        ticket.get("Nombre Sede Educativa A Beneficiar"),
        ticket.get("Código Dane Sede Educativa A Beneficiar"),
        ticket.get("Tipo Solución Engergía"),
        fmt_date(ticket.get("Fecha Inicio")),
        fmt_date(ticket.get("Fecha Cierre")),
        ticket.get("Estado"),
        ticket.get("Prioridad"),
        ticket.get("Afecta Servicio"),
        ticket.get("Responsabilidad"),
        ticket.get("Categoría"),
        ticket.get("Tiempo Solución Días Háb"),
        ticket.get("Tiempo Solución Días Cal"),
        ticket.get("Tiempo Paradas Reloj"),
        ticket.get("Descripción Apertura"),
        ticket.get("Solución"),
    ]


def _sheet_tickets(ws, tickets: list, sheet_title: str) -> None:
    n_cols = len(TICKET_COLS)
    span = f"A1:{get_column_letter(n_cols)}1"
    title_row(ws, sheet_title, span)

    headers = [c[0] for c in TICKET_COLS]
    widths  = [c[1] for c in TICKET_COLS]
    ws.append(headers)
    hdr_style(ws, 2, C_MED)
    col_widths(ws, widths)

    for i, ticket in enumerate(tickets):
        ws.append(_get_ticket_row(ticket))
        r = ws.max_row
        estado = str(ticket.get("Estado") or "").strip().lower()
        if estado in ("cerrado", "closed", "resuelto"):
            fgc = C_GREEN if i % 2 == 0 else C_GREEN2
        else:
            fgc = C_YELLOW if i % 2 == 0 else C_YELLOW2
        row_fill(ws, r, fgc, n_cols)

    freeze_filter(ws, "A3")


# ── Hoja 3: Paradas ───────────────────────────────────────────────────────────

PARADA_COLS = [
    ("Año-Mes",              12),
    ("Código Operador",      22),
    ("ID Beneficiario",      20),
    ("Departamento",         20),
    ("Municipio",            22),
    ("Centro Poblado",       26),
    ("Ticket #",             13),
    ("Fecha Inicio Falla",   18),
    ("Fecha Fin Falla",      18),
    ("Fecha Inicio Parada",  18),
    ("Fecha Fin Parada",     18),
    ("Horas Parada",         14),
    ("Días Parada",          12),
    ("Categoría Parada",     24),
]


def _sheet_paradas(ws, stops: list) -> None:
    n = len(PARADA_COLS)
    title_row(ws, "EVENTOS DE PARADA — DETALLE COMPLETO 2026", f"A1:{get_column_letter(n)}1")
    ws.append([c[0] for c in PARADA_COLS])
    hdr_style(ws, 2, C_MED)
    col_widths(ws, [c[1] for c in PARADA_COLS])

    for i, s in enumerate(stops):
        horas = s.get("Horas Parada Reloj")
        dias  = hours_to_days(horas)
        ws.append([
            s.get("Año-Mes"),
            s.get("Código Operador"),
            s.get("ID Beneficiario Mintic"),
            s.get("Departamento"),
            s.get("Municipio"),
            s.get("Centro Poblado"),
            s.get("Ticket #"),
            fmt_date(s.get("Fecha Inicio Falla")),
            fmt_date(s.get("Fecha Fin Falla")),
            fmt_date(s.get("Fecha Inicio Parada")),
            fmt_date(s.get("Fecha Fin Parada")),
            horas,
            dias,
            s.get("Categoría Parada Reloj"),
        ])
        row_fill(ws, ws.max_row, C_ACCENT if i % 2 == 0 else C_WHITE, n)

    freeze_filter(ws, "A3")


# ── Hoja 4: Disponibilidad ────────────────────────────────────────────────────

DISP_COLS = [
    ("Operador",               22),
    ("Año",                     8),
    ("Mes",                     8),
    ("ID Beneficiario",        20),
    ("ID Sede",                14),
    ("Departamento",           20),
    ("Municipio",              22),
    ("Centro Poblado",         26),
    ("Sede Principal",         30),
    ("DANE Sede Ppal.",        16),
    ("Sede a Beneficiar",      30),
    ("DANE Sede Benef.",       16),
    ("Fecha Inicio Operación", 20),
    ("Tipo Sol. Energía",      18),
    ("Umbral (%)",             12),
    ("DDA",                    10),
    ("Indisponibilidad Juntas",18),
    ("Disponibilidad Esperada",20),
    ("Disponibilidad Mes",     18),
    ("Cumplimiento",           16),
    ("Estado Contractual",     20),
]

DISP_SRC_KEYS = [
    "Operador", "Anio", "Mes", "Id Beneficiario", "Id Sede",
    "Departamento", "Municipio", "Centro_poblado",
    "Nombre Sede Educativa Principal", "Código Dane Sede Educativa Principal",
    "Nombre Sede Educativa a Beneficiar", "Código Dane Sede Educativa a Beneficiar",
    "Fecha Inicio Operación", "Tipo Solución Energía", "Umbral", "DDA",
    "Indisponibilidad  Juntas", "Disponibilidad Esperada",
    "Disponibilidad Mes", "Cumplimiento indicador", "Estado Contractual",
]


def _sheet_disp(ws, disp: list) -> None:
    n = len(DISP_COLS)
    title_row(ws, "DISPONIBILIDAD POR JUNTAS — ENERO A ABRIL 2026", f"A1:{get_column_letter(n)}1")
    ws.append([c[0] for c in DISP_COLS])
    hdr_style(ws, 2, C_MED)
    col_widths(ws, [c[1] for c in DISP_COLS])

    if not disp:
        ws["A3"].value = "Sin datos disponibles"
        return

    for i, row_data in enumerate(disp):
        ws.append([row_data.get(k) for k in DISP_SRC_KEYS])
        r = ws.max_row
        cumpl = str(row_data.get("Cumplimiento indicador") or "").strip().lower()
        if any(x in cumpl for x in ("cumple", " si", "sí")):
            fgc = C_GREEN if i % 2 == 0 else C_GREEN2
        elif "no" in cumpl:
            fgc = C_RED if i % 2 == 0 else C_RED2
        else:
            fgc = C_ACCENT if i % 2 == 0 else C_WHITE
        row_fill(ws, r, fgc, n)

    freeze_filter(ws, "A3")


# ── Hoja 5: Por Operador ──────────────────────────────────────────────────────

def _sheet_operador(ws, tickets: list, stops: list, disp: list) -> None:
    title_row(ws, "RESUMEN POR OPERADOR — ENERO A ABRIL 2026", "A1:L1")

    OP = defaultdict(lambda: {
        "tickets": 0, "abiertos": 0, "cerrados": 0,
        "paradas": 0, "horas": 0.0,
        "depts": set(), "munis": set(), "cats": defaultdict(int),
        "cumple": 0, "no_cumple": 0,
    })

    for t in tickets:
        op = str(t.get("Código Operador") or "").strip()
        if not op:
            continue
        OP[op]["tickets"] += 1
        estado = str(t.get("Estado") or "").strip().lower()
        if estado in ("cerrado", "closed", "resuelto"):
            OP[op]["cerrados"] += 1
        else:
            OP[op]["abiertos"] += 1
        dept = str(t.get("Departamento") or "").strip()
        muni = str(t.get("Municipio") or "").strip()
        if dept:
            OP[op]["depts"].add(dept)
        if muni:
            OP[op]["munis"].add(muni)
        cat = str(t.get("Categoría") or "Sin categoría").strip()
        OP[op]["cats"][cat] += 1

    for s in stops:
        op = str(s.get("Código Operador") or "").strip()
        if not op:
            continue
        OP[op]["paradas"] += 1
        OP[op]["horas"] += float(s.get("Horas Parada Reloj") or 0)

    for d in disp:
        op = str(d.get("Operador") or "").strip()
        if not op:
            continue
        cumpl = str(d.get("Cumplimiento indicador") or "").strip().lower()
        if any(x in cumpl for x in ("cumple", " si", "sí")):
            OP[op]["cumple"] += 1
        elif "no" in cumpl:
            OP[op]["no_cumple"] += 1

    COLS = [
        ("Código Operador",       24),
        ("Total Tickets",         14),
        ("Abiertos",              12),
        ("Cerrados",              12),
        ("% Cierre",              10),
        ("Eventos Parada",        14),
        ("Horas Parada",          14),
        ("Días Parada",           12),
        ("Departamentos",         14),
        ("Municipios",            12),
        ("Categoría Principal",   26),
        ("Sedes Cumplen Disp.",   18),
        ("Sedes No Cumplen",      16),
    ]
    headers = [c[0] for c in COLS]
    ws.append(headers)
    hdr_style(ws, 2, C_MED)
    col_widths(ws, [c[1] for c in COLS])

    for i, (op, d) in enumerate(sorted(OP.items())):
        total = d["tickets"]
        pct   = f"{d['cerrados'] / total * 100:.1f}%" if total else "N/A"
        dias  = round(d["horas"] / 24, 2)
        cat_p = max(d["cats"], key=d["cats"].get) if d["cats"] else ""
        ws.append([
            op, total, d["abiertos"], d["cerrados"], pct,
            d["paradas"], round(d["horas"], 1), dias,
            len(d["depts"]), len(d["munis"]), cat_p,
            d["cumple"], d["no_cumple"],
        ])
        r = ws.max_row
        fgc = C_ACCENT if i % 2 == 0 else C_WHITE
        fill = PatternFill("solid", fgColor=fgc)
        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci > 1 else "left")

    freeze_filter(ws, "A3")


# ── Hoja 6: Por Departamento ──────────────────────────────────────────────────

def _sheet_departamento(ws, tickets: list, stops: list) -> None:
    title_row(ws, "RESUMEN POR DEPARTAMENTO — ENERO A ABRIL 2026", "A1:I1")

    DEPT = defaultdict(lambda: {
        "tickets": 0, "abiertos": 0,
        "paradas": 0, "horas": 0.0,
        "munis": set(), "ops": set(),
        "cats": defaultdict(int),
    })

    for t in tickets:
        dept = str(t.get("Departamento") or "Sin dato").strip()
        DEPT[dept]["tickets"] += 1
        if str(t.get("Estado") or "").strip().lower() not in ("cerrado", "closed", "resuelto"):
            DEPT[dept]["abiertos"] += 1
        muni = str(t.get("Municipio") or "").strip()
        if muni:
            DEPT[dept]["munis"].add(muni)
        op = str(t.get("Código Operador") or "").strip()
        if op:
            DEPT[dept]["ops"].add(op)
        cat = str(t.get("Categoría") or "Sin categoría").strip()
        DEPT[dept]["cats"][cat] += 1

    for s in stops:
        dept = str(s.get("Departamento") or "Sin dato").strip()
        DEPT[dept]["paradas"] += 1
        DEPT[dept]["horas"] += float(s.get("Horas Parada Reloj") or 0)

    COLS = [
        ("Departamento",          24),
        ("Total Tickets",         14),
        ("Abiertos",              12),
        ("Cerrados",              12),
        ("Municipios",            14),
        ("Operadores",            14),
        ("Categoría Principal",   26),
        ("Eventos Parada",        14),
        ("Días Parada",           12),
    ]
    headers = [c[0] for c in COLS]
    ws.append(headers)
    hdr_style(ws, 2, C_MED)
    col_widths(ws, [c[1] for c in COLS])

    for i, (dept, d) in enumerate(sorted(DEPT.items(), key=lambda x: -x[1]["tickets"])):
        cerr = d["tickets"] - d["abiertos"]
        cat_p = max(d["cats"], key=d["cats"].get) if d["cats"] else ""
        ws.append([
            dept, d["tickets"], d["abiertos"], cerr,
            len(d["munis"]), len(d["ops"]), cat_p,
            d["paradas"], round(d["horas"] / 24, 2),
        ])
        r = ws.max_row
        fgc = C_ACCENT if i % 2 == 0 else C_WHITE
        fill = PatternFill("solid", fgColor=fgc)
        for ci in range(1, len(COLS) + 1):
            cell = ws.cell(row=r, column=ci)
            cell.fill = fill
            cell.alignment = Alignment(vertical="center",
                                       horizontal="center" if ci > 1 else "left")

    freeze_filter(ws, "A3")


# ── Hoja 8: Cumplimiento SLA ──────────────────────────────────────────────────

def _sheet_sla(ws, tickets: list, stops: list) -> None:
    n_cols = 14
    title_row(ws, "ANÁLISIS DE CUMPLIMIENTO SLA — 2026", f"A1:{get_column_letter(n_cols)}1")

    stop_agg: dict[str, dict] = defaultdict(lambda: {"horas": 0.0, "eventos": 0})
    for s in stops:
        tid = str(s.get("Ticket #") or "").strip()
        if tid:
            stop_agg[tid]["horas"]   += float(s.get("Horas Parada Reloj") or 0)
            stop_agg[tid]["eventos"] += 1

    COLS = [
        ("#Ticket",             12),
        ("Mes",                 12),
        ("Código Operador",     22),
        ("Departamento",        20),
        ("Municipio",           22),
        ("Estado",              13),
        ("Prioridad",           13),
        ("Fecha Inicio",        17),
        ("Fecha Cierre",        17),
        ("Días Háb. Sol.",      14),
        ("Días Cal. Sol.",      14),
        ("Horas Parada",        14),
        ("Días Parada",         12),
        ("Estado SLA",          16),
    ]
    ws.append([c[0] for c in COLS])
    hdr_style(ws, 2, C_MED)
    col_widths(ws, [c[1] for c in COLS])

    SLA_DIAS_HAB = 5

    for i, t in enumerate(tickets):
        tid        = str(t.get("#Ticket") or "").strip()
        dias_hab   = t.get("Tiempo Solución Días Háb")
        try:
            dias_hab_f = float(dias_hab) if dias_hab is not None else None
        except (TypeError, ValueError):
            dias_hab_f = None

        sa         = stop_agg.get(tid, {})
        horas_p    = sa.get("horas", 0.0)
        dias_p     = round(horas_p / 24, 2)
        estado_raw = str(t.get("Estado") or "").strip().lower()

        if dias_hab_f is not None and dias_hab_f > SLA_DIAS_HAB:
            sla = "EXCEDE SLA"
            fgc = C_RED if i % 2 == 0 else C_RED2
        elif dias_p > 7:
            sla = "PARADA CRÍTICA"
            fgc = C_YELLOW if i % 2 == 0 else C_YELLOW2
        elif estado_raw in ("cerrado", "closed", "resuelto"):
            sla = "Cerrado OK"
            fgc = C_GREEN if i % 2 == 0 else C_GREEN2
        else:
            sla = "Abierto"
            fgc = C_ACCENT if i % 2 == 0 else C_WHITE

        ws.append([
            tid,
            t.get("__mes__"),
            t.get("Código Operador"),
            t.get("Departamento"),
            t.get("Municipio"),
            t.get("Estado"),
            t.get("Prioridad"),
            fmt_date(t.get("Fecha Inicio")),
            fmt_date(t.get("Fecha Cierre")),
            t.get("Tiempo Solución Días Háb"),
            t.get("Tiempo Solución Días Cal"),
            round(horas_p, 1) if horas_p else None,
            dias_p if dias_p else None,
            sla,
        ])
        row_fill(ws, ws.max_row, fgc, n_cols)

    freeze_filter(ws, "A3")


if __name__ == "__main__":
    main()
